from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import Cliente, Proveedor, Inventario, HistorialProveedoresNotas, MovimientosInventario, PerfilUsuario, PedidoCompra, DetallePedidoCompra
from .forms import ClienteForm, ProveedorForm, InventarioForm, HistorialProveedoresNotasForm, MovimientosInventarioForm, UserForm, PerfilUsuarioForm, ImportarArchivoForm
from django.core.paginator import Paginator
from django.db.models import Q, F, Sum, Count, Value, Max
from django.utils import timezone
from django.db import transaction
from datetime import timedelta, datetime
from django.shortcuts import get_object_or_404
from django import forms
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from fpdf import FPDF
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth import logout
from django.core.management import call_command
import uuid
import openpyxl
import csv
import io
import uuid
# Create your views here.
# libreria/views.py
 
import json
# ...
def custom_logout(request):
    logout(request)
    return redirect('login')

# --- Decorador para verificar rol de Administrador ---
def es_admin(user):
    return hasattr(user, 'perfilusuario') and user.perfilusuario.rol == 'admin'

def es_inventario_acceso(user):
    return hasattr(user, 'perfilusuario') and user.perfilusuario.rol in ['admin', 'inventario', 'consulta']

# Nuevo decorador para restringir acceso a Proveedores y Productos (excluye a 'consulta')
def es_pleno_acceso(user):
    return hasattr(user, 'perfilusuario') and user.perfilusuario.rol in ['admin', 'inventario']

@login_required
def index(request):
    # Si el usuario es rol 'consulta', redirigir directo a Movimientos SOLO la primera vez (al iniciar sesión)
    if hasattr(request.user, 'perfilusuario') and request.user.perfilusuario.rol == 'consulta':
        if not request.session.get('consulta_redirected'):
            request.session['consulta_redirected'] = True
            return redirect('movimientos.index')

    # --- CÁLCULO PARA LAS TARJETAS (CARDS) ---

    # 1. Total de proveedores
    total_proveedores = Proveedor.objects.count()

    # 2. Total de productos
    total_productos = Inventario.objects.count()

    # 4. Producto con más movimientos (ENTRADA y SALIDA)
    pm = MovimientosInventario.objects.exclude(tipo_movimiento='PEDIDO').values(
        'producto__nombre_producto',
        'producto__stock_minimo',
        'producto__stock_maximo',
        'producto__cantidad'
    ).annotate(
        total_movimientos=Sum('cantidad')
    ).order_by('-total_movimientos').first()

    producto_top = {
        'nombre': pm['producto__nombre_producto'] if pm else "N/A",
        'total': pm['total_movimientos'] if pm else 0,
        'min': pm['producto__stock_minimo'] if pm else 0,
        'max': pm['producto__stock_maximo'] if pm else 0,
        'stock': pm['producto__cantidad'] if pm else 0,
    }

    # 5. LISTADO DE ÚLTIMOS 5 MOVIMIENTOS (NUEVO)
    ultimos_movimientos = MovimientosInventario.objects.exclude(tipo_movimiento='PEDIDO').select_related('producto', 'proveedor').order_by('-fecha_movimiento')[:5]

    # 6. Alerta de Stock Bajo
    productos_bajo_stock = Inventario.objects.filter(
        cantidad__lte=F('stock_minimo')
    ).values('nombre_producto', 'cantidad', 'stock_minimo')

    # --- CONTEXTO PARA LA PLANTILLA ---
    context = {
        'producto_top': producto_top,
        'ultimos_movimientos': ultimos_movimientos,
        'productos_bajo_stock': productos_bajo_stock,
        'num_bajo_stock': productos_bajo_stock.count(),
    }

    return render(request, 'index.html', context)


#======================================
#proveedores vista
#========================================

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def proveedores_index(request):
    # parámetros GET
    q = request.GET.get('q', '').strip()
    order = request.GET.get('order', 'razonsocial_asc')
    page_size = 10

    # queryset base
    qs = Proveedor.objects.all()

    # búsqueda por nombre, rif o razón social (insensible a mayúsc/minúsc)
    if q:
        qs = qs.filter(
            Q(nombre__icontains=q) |
            Q(rif__icontains=q) |
            Q(razonsocial__icontains=q)
        )

    # orden opcional
    if order == 'razonsocial_desc':
        qs = qs.order_by('-razonsocial')
    else:
        qs = qs.order_by('razonsocial')

    # paginación
    paginator = Paginator(qs, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'proveedores/index.html', {
        'page_obj': page_obj,
        'q': q,
        'order': order,
    })

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def proveedores_crear(request):
    formulario_proveedores = ProveedorForm(request.POST or None)
    if formulario_proveedores.is_valid():
        formulario_proveedores.save()
        messages.success(request, 'Proveedor registrado exitosamente.')
        return redirect('/proveedores')
    return render(request, 'proveedores/crear.html', {'formulario_proveedores': formulario_proveedores})

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def proveedores_editar(request,id):
    proveedor = Proveedor.objects.get(id=id)
    formulario_proveedores = ProveedorForm(request.POST or None, instance=proveedor)
    if formulario_proveedores.is_valid() and request.POST:
        formulario_proveedores.save()
        messages.success(request, 'Proveedor actualizado exitosamente.')
        return redirect('/proveedores')
    return render(request, 'proveedores/editar.html',{'formulario_proveedores': formulario_proveedores})

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def proveedores_eliminar(request,id):
    proveedores = Proveedor.objects.get(id=id)
    proveedores.delete()
    messages.success(request, 'Proveedor eliminado exitosamente.')
    return redirect('/proveedores')

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def proveedores_importar(request):
    if request.method == 'POST':
        form = ImportarArchivoForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            nombre_archivo = archivo.name.lower()
            registros_creados = 0
            errores = []

            try:
                # Determinar si es Excel o CSV
                datos = []
                if nombre_archivo.endswith('.xlsx'):
                    wb = openpyxl.load_workbook(archivo)
                    ws = wb.active
                    # Iterar filas, saltando la cabecera (min_row=2)
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        # Esperamos: Nombre, Telefono, Razon Social, RIF, Direccion
                        if row[0]: # Si hay nombre
                            datos.append(row)
                elif nombre_archivo.endswith('.csv') or nombre_archivo.endswith('.txt'):
                    archivo_data = archivo.read().decode('utf-8')
                    io_string = io.StringIO(archivo_data)
                    reader = csv.reader(io_string, delimiter=',')
                    next(reader, None) # Saltar cabecera
                    for row in reader:
                        if row and row[0]:
                            datos.append(row)
                
                # Procesar datos
                for i, fila in enumerate(datos):
                    try:
                        # Asumiendo orden: Nombre, Telefono, Razon Social, RIF, Direccion
                        Proveedor.objects.create(
                            nombre=fila[0],
                            telefono=fila[1] if len(fila) > 1 else '',
                            razonsocial=fila[2] if len(fila) > 2 else '',
                            rif=fila[3] if len(fila) > 3 else f"S/R-{i}", # RIF es unique, cuidado
                            direccion=fila[4] if len(fila) > 4 else ''
                        )
                        registros_creados += 1
                    except Exception as e:
                        errores.append(f"Fila {i+2}: {str(e)}")

                messages.success(request, f'Se importaron {registros_creados} proveedores correctamente.')
                if errores:
                    messages.warning(request, f'Hubo errores en {len(errores)} filas. Primera falla: {errores[0]}')
                return redirect('proveedores.index')

            except Exception as e:
                messages.error(request, f'Error procesando el archivo: {str(e)}')

    else:
        form = ImportarArchivoForm()

    return render(request, 'proveedores/importar.html', {'form': form})

#======================================
#inventario vista
#========================================

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def inventario_index(request):
     # Query base
    qs = Inventario.objects.all()

    # Parámetros GET
    q = request.GET.get('q', '').strip()
    categoria = request.GET.get('categoria', '').strip()
    proveedor_id = request.GET.get('proveedor_id', '').strip()
    order = request.GET.get('order', 'name_asc')  # 'name_asc','name_desc','cantidad_desc','cantidad_asc'
    low = request.GET.get('low', '').strip()  # si '1' o 'true' -> filtrar los bajos
    page_size = 10

    # Filtro por texto en nombre, descripción o código
    if q:
        qs = qs.filter(
            Q(nombre_producto__icontains=q) |
            Q(codigo_producto__icontains=q)
        )
 
    # Filtrar por categoría si se pasa
    if categoria:
        qs = qs.filter(categoria=categoria)
        
    # Filtrar por proveedor_id si se pasa
    if proveedor_id:
        qs = qs.filter(proveedores__id=proveedor_id)

    # Filtrar sólo productos con stock bajo si se solicita
    if low and low.lower() in ('1', 'true', 'on'):
        qs = qs.filter(cantidad__lte=F('stock_minimo'))

    # Orden
    if order == 'name_desc':
        qs = qs.order_by('-nombre_producto')
    elif order == 'cantidad_desc':
        qs = qs.order_by('-cantidad')
    elif order == 'cantidad_asc':
        qs = qs.order_by('cantidad')
    else:
        qs = qs.order_by('nombre_producto')

    # Contador de productos con stock bajo (sobre el conjunto ya filtrado)
    low_stock_count = qs.filter(cantidad__lte=F('stock_minimo')).count()

    # Paginación
    paginator = Paginator(qs, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'q': q,
        'categoria': categoria,
        'order': order,
        'low_stock_count': low_stock_count,
        'low': low,
        'proveedor_id': proveedor_id,
    }
    return render(request, 'inventario/index.html', context)

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def inventario_crear(request):
    formulario_inventario = InventarioForm(request.POST or None)
    if formulario_inventario.is_valid():
        formulario_inventario.save()
        messages.success(request, 'Producto registrado exitosamente.')
        return redirect('/inventario')
    return render(request, 'inventario/crear.html',{'formulario_inventario': formulario_inventario})

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def inventario_editar(request,id_producto):
    producto = Inventario.objects.get(id_producto=id_producto)
    formulario_inventario = InventarioForm(request.POST or None, instance=producto)
    if formulario_inventario.is_valid() and request.POST:
        formulario_inventario.save()
        messages.success(request, 'Producto actualizado exitosamente.')
        return redirect('/inventario')
    return render(request, 'inventario/editar.html',{'formulario_inventario': formulario_inventario})

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def inventario_eliminar(request,id_producto):
    producto = Inventario.objects.get(id_producto=id_producto)
    producto.delete()
    messages.success(request, 'Producto eliminado exitosamente.')
    return redirect('/inventario')

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def inventario_importar(request):
    if request.method == 'POST':
        form = ImportarArchivoForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            nombre_archivo = archivo.name.lower()
            registros_creados = 0
            errores = []

            try:
                datos = []
                if nombre_archivo.endswith('.xlsx'):
                    wb = openpyxl.load_workbook(archivo)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]: # Si hay código o nombre
                            datos.append(row)
                elif nombre_archivo.endswith('.csv') or nombre_archivo.endswith('.txt'):
                    archivo_data = archivo.read().decode('utf-8')
                    io_string = io.StringIO(archivo_data)
                    reader = csv.reader(io_string, delimiter=',')
                    next(reader, None)
                    for row in reader:
                        if row:
                            datos.append(row)

                for i, fila in enumerate(datos):
                    try:
                        # Orden esperado: Codigo, Nombre, Descripcion, Categoria, Cantidad, Costo, StockMin, StockMax
                        # Mapeo simple de categoría (si no coincide, usa OTRO)
                        cat_input = str(fila[3]).upper().strip() if len(fila) > 3 else 'OTRO'
                        categoria_valida = 'OTRO'
                        for choice in Inventario._meta.get_field('categoria').choices:
                            if choice[0] == cat_input or choice[1].upper() == cat_input:
                                categoria_valida = choice[0]
                                break
                        
                        # Ahora busca si el producto ya existe por su código
                        obj, created = Inventario.objects.update_or_create(
                            codigo_producto=fila[0],
                            defaults={
                                'nombre_producto': fila[1] if len(fila) > 1 else 'Sin Nombre',
                                'descripcion': fila[2] if len(fila) > 2 else '',
                                'categoria': categoria_valida,
                                'cantidad': int(fila[4]) if len(fila) > 4 and fila[4] else 0,
                                'costo_actual': float(fila[5]) if len(fila) > 5 and fila[5] else 0.0,
                                'stock_minimo': int(fila[6]) if len(fila) > 6 and fila[6] else 5,
                                'stock_maximo': int(fila[7]) if len(fila) > 7 and fila[7] else 100,
                                'unidad_empaque': 'UNIDAD',
                                'cantidad_por_empaque': 1
                            }
                        )

                        # Si hay una 9na columna, intenta asignar el proveedor
                        if len(fila) > 8 and fila[8]:
                            valor_proveedor = str(fila[8]).strip()
                            # Busca por Nombre o por RIF
                            proveedor = Proveedor.objects.filter(
                                Q(nombre__icontains=valor_proveedor) | Q(rif__icontains=valor_proveedor)
                            ).first()
                            if proveedor:
                                obj.proveedores.add(proveedor)
                        registros_creados += 1
                    except Exception as e:
                        errores.append(f"Fila {i+2} ({fila[0] if fila else '?'}): {str(e)}")

                messages.success(request, f'Se importaron {registros_creados} productos.')
                if errores:
                    messages.warning(request, f'Errores: {len(errores)}. {errores[0]}')
                return redirect('inventario.index')
            except Exception as e:
                messages.error(request, f'Error general: {str(e)}')
    else:
        form = ImportarArchivoForm()
    return render(request, 'inventario/importar.html', {'form': form})

# --- EXPORTACIÓN DE INFORMES ---

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def exportar_inventario_excel(request):
    # Crear un libro de trabajo (workbook)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Encabezados
    headers = ["Código", "Producto", "Categoría", "Cantidad", "Unidad Empaque", "Costo Actual", "Valor Total"]
    ws.append(headers)

    # Obtener datos (puedes aplicar los mismos filtros que en el index si quisieras, aquí exportamos todo)
    productos = Inventario.objects.all()

    for p in productos:
        valor_total = p.cantidad * p.costo_actual
        ws.append([
            p.codigo_producto,
            p.nombre_producto,
            p.get_categoria_display(),
            p.cantidad,
            p.get_unidad_empaque_display(),
            p.costo_actual,
            valor_total
        ])

    # Preparar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario_reporte.xlsx"'
    
    wb.save(response)
    return response

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def exportar_inventario_pdf(request):
    # Obtener datos
    productos = Inventario.objects.all().order_by('nombre_producto')
    
    # Calcular totales generales para el reporte
    total_items = productos.count()
    valor_total_inventario = sum(p.cantidad * p.costo_actual for p in productos)
    fecha_emision = timezone.now().strftime('%d/%m/%Y %H:%M')

    # Crear PDF usando fpdf2 (Pure Python)
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    
    # Título
    pdf.cell(0, 10, "Reporte de Inventario General", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 10, f"Fecha de Emision: {fecha_emision}", ln=True, align="C")
    pdf.ln(5)

    # Resumen
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 10, f"Resumen General:", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 7, f"Total de Productos: {total_items}", ln=True)
    pdf.cell(0, 7, f"Valor Total del Inventario: {valor_total_inventario:,.2f}", ln=True)
    pdf.ln(5)

    # Tabla de Productos
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(240, 240, 240)
    
    # Encabezados
    pdf.cell(30, 10, "Codigo", 1, 0, "C", True)
    pdf.cell(70, 10, "Producto", 1, 0, "C", True)
    pdf.cell(30, 10, "Categoria", 1, 0, "C", True)
    pdf.cell(20, 10, "Stock", 1, 0, "C", True)
    pdf.cell(40, 10, "Valor Total", 1, 1, "C", True)

    pdf.set_font("Helvetica", "", 9)
    for p in productos:
        valor_total = p.cantidad * p.costo_actual
        
        # Truncar nombres largos para que no se desborden
        nombre = p.nombre_producto[:35]
        categoria = p.get_categoria_display()[:15]
        
        pdf.cell(30, 8, str(p.codigo_producto), 1, 0, "L")
        pdf.cell(70, 8, nombre, 1, 0, "L")
        pdf.cell(30, 8, categoria, 1, 0, "L")
        pdf.cell(20, 8, str(p.cantidad), 1, 0, "C")
        pdf.cell(40, 8, f"{valor_total:,.2f}", 1, 1, "R")

    # Preparar la respuesta HTTP
    response = HttpResponse(pdf.output(dest='S').encode('latin-1'), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="inventario_reporte.pdf"'
    
    return response

#======================================
# notas vista
#========================================

@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def historial_proveedores_notas_index(request):
    # Query base: traer todos los movimientos de inventario
    base_qs = MovimientosInventario.objects.all()

    # Get params
    q = request.GET.get('q', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    order = request.GET.get('order', 'desc')

    # Filtro por texto en producto, proveedor o tipo de movimiento
    if q:
        base_qs = base_qs.filter(
            Q(producto__nombre_producto__icontains=q) |
            Q(proveedor__razonsocial__icontains=q) |
            Q(tipo_movimiento__icontains=q)
        )

    # Filtro por rango de fechas
    if date_from:
        try:
            fecha_inicio = datetime.strptime(date_from, '%Y-%m-%d')
            base_qs = base_qs.filter(fecha_movimiento__gte=fecha_inicio)
        except ValueError:
            pass
    if date_to:
        try:
            fecha_fin = datetime.strptime(date_to, '%Y-%m-%d')
            fecha_fin = fecha_fin.replace(hour=23, minute=59, second=59)
            base_qs = base_qs.filter(fecha_movimiento__lte=fecha_fin)
        except ValueError:
            pass

    # Agrupación y anotación
    qs = base_qs.values(
        'proveedor__razonsocial',
        'proveedor__nombre',
        'producto__nombre_producto',
        'tipo_movimiento'
    ).annotate(
        total_unidades=Sum('cantidad'),
        total_empaques=Sum('cantidad_empaques'),
        ultima_fecha=Max('fecha_movimiento')
    )

    # Orden
    if order == 'asc':
        qs = qs.order_by('ultima_fecha')
    else:
        qs = qs.order_by('-ultima_fecha')

    # Paginación
    page_size = 10
    paginator = Paginator(qs, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'q': q,
        'date_from': date_from,
        'date_to': date_to,
        'order': order,
    }
    return render(request, 'HistorialProveedoresNotas/index.html', context)


@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def historial_proveedores_notas_crear(request):
    formulario_nota = HistorialProveedoresNotasForm(request.POST or None)

    if formulario_nota.is_valid():
        formulario_nota.save()
        messages.success(request, 'Nota registrada exitosamente.')
        return redirect('HistorialProveedoresNotas.index')
    return render(request, 'HistorialProveedoresNotas/crear.html', {'formulario_nota': formulario_nota})

@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def historial_proveedores_notas_editar(request,id_historialproveedor):
    nota = HistorialProveedoresNotas.objects.get(id_historialproveedor=id_historialproveedor)
    formulario_nota = HistorialProveedoresNotasForm(request.POST or None, instance=nota)

    if formulario_nota.is_valid() and request.POST:
        formulario_nota.save()
        messages.success(request, 'Nota actualizada exitosamente.')
        return redirect('HistorialProveedoresNotas.index')
    return render(request, 'HistorialProveedoresNotas/editar.html', {
        'formulario_nota': formulario_nota,
    })

@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def historial_proveedores_notas_eliminar(request,id_historialproveedor):
    nota = HistorialProveedoresNotas.objects.get(id_historialproveedor=id_historialproveedor)
    nota.delete()
    messages.success(request, 'Nota eliminada exitosamente.')
    return redirect('HistorialProveedoresNotas.index')

#======================================
#MovimientosInventario vista
#========================================

@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def movimientos_inventario_index(request):
    # Parámetros GET para filtros y orden
    q = request.GET.get('q', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    order = request.GET.get('order', 'desc')
    page_size = 10

    # Queryset base con optimización para evitar consultas N+1
    qs = MovimientosInventario.objects.select_related('producto', 'proveedor')

    # Filtro de búsqueda por nombre de producto o proveedor
    if q:
        qs = qs.filter(
            Q(producto__nombre_producto__icontains=q) |
            Q(proveedor__razonsocial__icontains=q) |
            Q(proveedor__rif__icontains=q)
        )

    # Filtro por rango de fechas
    if date_from:
        try:
            # Convertimos a datetime (00:00:00 del día seleccionado)
            fecha_inicio = datetime.strptime(date_from, '%Y-%m-%d')
            qs = qs.filter(fecha_movimiento__gte=fecha_inicio)
        except ValueError:
            pass
    if date_to:
        try:
            # Convertimos a datetime y ajustamos al final del día (23:59:59)
            fecha_fin = datetime.strptime(date_to, '%Y-%m-%d')
            fecha_fin = fecha_fin.replace(hour=23, minute=59, second=59)
            qs = qs.filter(fecha_movimiento__lte=fecha_fin)
        except ValueError:
            pass

    # Calcular totales solo si hay un filtro de búsqueda activo
    resumen_filtro = None
    if q:
        # 1. Agrupar por producto y calcular totales
        resumen_por_producto = qs.values('producto__nombre_producto') \
                                 .annotate(
                                     entradas=Sum('cantidad', default=0, filter=Q(tipo_movimiento='ENTRADA')),
                                     salidas=Sum('cantidad', default=0, filter=Q(tipo_movimiento='SALIDA'))
                                 ).order_by('producto__nombre_producto')
        
        # 2. Construir las cadenas de texto para la vista y calcular totales generales
        entradas_str_parts = []
        salidas_str_parts = []
        total_entradas_general = 0
        total_salidas_general = 0

        for item in resumen_por_producto:
            if item['entradas'] and item['entradas'] > 0:
                entradas_str_parts.append(f"{item['entradas']} ({item['producto__nombre_producto']})")
                total_entradas_general += item['entradas']
            if item['salidas'] and item['salidas'] > 0:
                salidas_str_parts.append(f"{item['salidas']} ({item['producto__nombre_producto']})")
                total_salidas_general += item['salidas']

        # 3. Calcular el stock actual de los productos encontrados
        ids_productos = qs.values_list('producto_id', flat=True).distinct()
        
        # Obtenemos el detalle del stock por producto para mostrarlo desglosado
        productos_stock = Inventario.objects.filter(id_producto__in=ids_productos).values('nombre_producto', 'cantidad').order_by('nombre_producto')
        
        stock_desglose = []
        total_stock_general = 0
        
        for p in productos_stock:
            stock_desglose.append(f"{p['cantidad']} ({p['nombre_producto']})")
            total_stock_general += p['cantidad']

        # 4. Ensamblar el diccionario final para el contexto
        resumen_filtro = {
            'entradas_desglose': entradas_str_parts,
            'salidas_desglose': salidas_str_parts,
            'stock_desglose': stock_desglose,
            'total_entradas_general': total_entradas_general,
            'total_salidas_general': total_salidas_general,
            'stock_actual': total_stock_general
        }
    # Ordenamiento: Si hay una búsqueda, agrupa por producto.
    if q:
        # Cuando se busca, se agrupa por producto y luego por fecha para claridad.
        if order == 'asc':
            qs = qs.order_by('producto__nombre_producto', 'fecha_movimiento')
        else: # 'desc' es el default
            qs = qs.order_by('producto__nombre_producto', '-fecha_movimiento')
    else:
        # Comportamiento original si no hay búsqueda (solo por fecha).
        if order == 'asc':
            qs = qs.order_by('fecha_movimiento')
        else:
            qs = qs.order_by('-fecha_movimiento')

    # Paginación
    paginator = Paginator(qs, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj, 'q': q, 'order': order,
        'resumen_filtro': resumen_filtro,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'movimientos/index.html', context)

@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def movimientos_inventario_crear(request):
    formulario = MovimientosInventarioForm(request.POST or None)
    
    if formulario.is_valid():
        formulario.save()
        messages.success(request, 'Movimiento registrado y stock actualizado correctamente.')
        return redirect('movimientos.index')

    return render(request, 'movimientos/crear.html', {'formulario': formulario})


@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def movimientos_inventario_editar(request, id_movimiento):
    movimiento = get_object_or_404(MovimientosInventario, id_movimiento=id_movimiento)
    
    if request.method == 'POST':
        formulario = MovimientosInventarioForm(request.POST, instance=movimiento)
        if formulario.is_valid():
            # El método save del formulario maneja toda la lógica de stock y transacciones.
            # Puede lanzar ValidationError si el stock es insuficiente durante la edición.
            try:
                formulario.save()
                messages.success(request, 'El movimiento se ha actualizado y el stock ha sido ajustado correctamente.')
                return redirect('movimientos.index')
            except forms.ValidationError as e:
                # Agrega el error al formulario para que se muestre en la plantilla
                formulario.add_error(None, e)
    else:
        formulario = MovimientosInventarioForm(instance=movimiento)

    return render(request, 'movimientos/editar.html', {
        'formulario': formulario,
        'movimiento': movimiento
    })


@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def movimientos_inventario_eliminar(request, id_movimiento):
    # Usamos select_related para traer el producto en la misma consulta
    movimiento = MovimientosInventario.objects.select_related('producto').get(id_movimiento=id_movimiento)
    
    if request.method == 'POST':
        with transaction.atomic():
            producto = movimiento.producto
            # Revertir el stock antes de eliminar el movimiento
            if movimiento.tipo_movimiento == 'ENTRADA':
                producto.cantidad -= movimiento.cantidad
            elif movimiento.tipo_movimiento == 'SALIDA':
                producto.cantidad += movimiento.cantidad
            
            producto.save()
            movimiento.delete()
        messages.success(request, 'Movimiento eliminado exitosamente.')
        return redirect('movimientos.index')
    
    return render(request, 'movimientos/eliminar.html', {'movimiento': movimiento})


@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def movimientos_salida_form(request):
    from .models import UnidadEmpaqueChoices
    qs = Inventario.objects.all().order_by('nombre_producto')
    unidades_choices = UnidadEmpaqueChoices.choices
    
    # Pre-renderizar las opciones de unidades para ahorrar JS en el front
    unidades_html = "".join([f'<option value="{v}">{l}</option>' for v, l in unidades_choices])

    # Preparar datos para el buscador en el frontend
    productos_json = []
    for p in qs:
        productos_json.append({
            'id': p.id_producto,
            'nombre': p.nombre_producto,
            'codigo': p.codigo_producto,
            'stock': p.cantidad,
            'unidades_html': unidades_html,
            'id_unidad_default': p.unidad_empaque
        })

    return render(request, 'movimientos/form_salida.html', {
        'productos_json': productos_json,
        'unidades_choices': unidades_choices,
    })


@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def movimientos_salida_confirmar(request):
    if request.method == 'POST':
        producto_ids = request.POST.getlist('producto_id[]')
        unidades_empaque = request.POST.getlist('unidad_empaque[]')
        cants_empaques = request.POST.getlist('cant_empaques[]')
        totales = request.POST.getlist('total_unidades[]')
        
        # Filtramos solo los que tienen cantidad > 0
        items_resumen = []
        for i in range(len(producto_ids)):
            cant = int(totales[i])
            if cant > 0:
                producto = Inventario.objects.get(id_producto=producto_ids[i])
                items_resumen.append({
                    'producto': producto,
                    'unidad': unidades_empaque[i],
                    'cant_empaques': cants_empaques[i],
                    'total': cant,
                })
        
        if not items_resumen:
            messages.warning(request, "Debe seleccionar al menos un producto con cantidad mayor a cero.")
            return redirect('movimientos.salida')
            
        proveedores = Proveedor.objects.all().order_by('razonsocial')
        
        return render(request, 'movimientos/confirmar_salida.html', {
            'items': items_resumen,
            'proveedores': proveedores
        })
    return redirect('movimientos.index')

@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def movimientos_salida_procesar(request):
    if request.method == 'POST':
        producto_ids = request.POST.getlist('producto_id[]')
        unidades_empaque = request.POST.getlist('unidad_empaque[]')
        cants_empaques = request.POST.getlist('cant_empaques[]')
        totales = request.POST.getlist('total_unidades[]')
        proveedor_id = request.POST.get('proveedor_id')

        if not proveedor_id:
            messages.error(request, 'Debe seleccionar un proveedor para el descargo.')
            return redirect('movimientos.salida')

        proveedor = Proveedor.objects.get(id=proveedor_id)
        ahora = timezone.now()
        lote_id = f"S-{ahora.strftime('%Y%m%d%H%M')}-{str(uuid.uuid4())[:8]}"
        salidas_creadas = 0
        try:
            with transaction.atomic():
                for i in range(len(producto_ids)):
                    producto = Inventario.objects.get(id_producto=producto_ids[i])
                    cant_salida = int(totales[i])
                    
                    if cant_salida > 0:
                        # VALIDACIÓN DE STOCK
                        if producto.cantidad < cant_salida:
                            raise forms.ValidationError(f"Stock insuficiente para {producto.nombre_producto}. Disponible: {producto.cantidad}")

                        # Crear el movimiento
                        MovimientosInventario.objects.create(
                            producto=producto,
                            tipo_movimiento='SALIDA',
                            cantidad=cant_salida,
                            unidad_empaque=unidades_empaque[i],
                            cantidad_empaques=int(cants_empaques[i]),
                            proveedor=proveedor,
                            fecha_movimiento=ahora,
                            codigo_lote=lote_id
                        )
                        # Actualizar stock
                        producto.cantidad -= cant_salida
                        producto.save()
                        salidas_creadas += 1
            
            messages.success(request, f'Se registraron exitosamente {salidas_creadas} salidas del inventario con descargo a {proveedor.razonsocial}.')
            return redirect('movimientos.index')
            
        except forms.ValidationError as e:
            messages.error(request, str(e))
            return redirect('movimientos.salida')
        except Exception as e:
            messages.error(request, f'Error al procesar las salidas: {str(e)}')
            return redirect('movimientos.salida')

    return redirect('movimientos.index')

@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def movimientos_entrada_form(request, pedido_id=None):
    from .models import UnidadEmpaqueChoices
    unidades_choices = UnidadEmpaqueChoices.choices
    unidades_html = "".join([f'<option value="{v}">{l}</option>' for v, l in unidades_choices])

    pedido = None
    items_pedido = []
    if pedido_id:
        pedido = get_object_or_404(PedidoCompra, id_pedido=pedido_id)
        # Pre-cargar items del pedido
        for detalle in pedido.detalles.all():
            items_pedido.append({
                'id': detalle.producto.id_producto,
                'nombre': detalle.producto.nombre_producto,
                'codigo': detalle.producto.codigo_producto,
                'cantidad': detalle.cantidad,
                'unidad': detalle.unidad_empaque,
                'empaques': detalle.cantidad_empaques,
                'por_empaque': detalle.cantidad_por_empaque
            })
    
    # Preparar productos para el buscador (si quieren agregar nuevos)
    qs = Inventario.objects.all().order_by('nombre_producto')
    productos_json = []
    for p in qs:
        productos_json.append({
            'id': p.id_producto,
            'nombre': p.nombre_producto,
            'codigo': p.codigo_producto,
            'unidades_html': unidades_html,
            'id_unidad_default': p.unidad_empaque
        })

    proveedores = Proveedor.objects.all().order_by('razonsocial')
    
    # Preparar listado de pedidos informativos para búsqueda rápida
    pedidos_qs = PedidoCompra.objects.select_related('proveedor').prefetch_related('detalles__producto').order_by('-fecha_pedido')[:50]
    pedidos_json = []
    for ped in pedidos_qs:
        items = []
        for det in ped.detalles.all():
            items.append({
                'id': det.producto.id_producto,
                'nombre': det.producto.nombre_producto,
                'codigo': det.producto.codigo_producto,
                'cantidad': det.cantidad,
                'unidad': det.unidad_empaque,
                'empaques': det.cantidad_empaques,
                'por_empaque': det.cantidad_por_empaque
            })
        pedidos_json.append({
            'id': ped.id_pedido,
            'codigo': ped.codigo_lote,
            'proveedor_id': ped.proveedor.id,
            'proveedor_nombre': ped.proveedor.razonsocial,
            'fecha': ped.fecha_pedido.strftime('%d/%m/%Y'),
            'items': items
        })

    return render(request, 'movimientos/form_entrada.html', {
        'productos_json': productos_json,
        'unidades_choices': unidades_choices,
        'pedido': pedido,
        'items_pedido': items_pedido,
        'proveedores': proveedores,
        'pedidos_json': pedidos_json, # Nueva data para búsqueda dinámica
    })

@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def movimientos_entrada_confirmar(request):
    if request.method == 'POST':
        producto_ids = request.POST.getlist('producto_id[]')
        unidades_empaque = request.POST.getlist('unidad_empaque[]')
        cants_empaques = request.POST.getlist('cant_empaques[]')
        totales = request.POST.getlist('total_unidades[]')
        proveedor_id = request.POST.get('proveedor_id')
        
        items_resumen = []
        for i in range(len(producto_ids)):
            cant = int(totales[i])
            if cant > 0:
                producto = Inventario.objects.get(id_producto=producto_ids[i])
                items_resumen.append({
                    'producto': producto,
                    'unidad': unidades_empaque[i],
                    'cant_empaques': cants_empaques[i],
                    'total': cant,
                })
        
        if not items_resumen:
            messages.warning(request, "Debe seleccionar al menos un producto con cantidad mayor a cero.")
            return redirect('movimientos.entrada')
            
        pedido_id_vincular = request.POST.get('pedido_id_vincular')
            
        proveedor = None
        if proveedor_id:
            proveedor = get_object_or_404(Proveedor, id=proveedor_id)
        
        return render(request, 'movimientos/confirmar_entrada.html', {
            'items': items_resumen,
            'proveedor': proveedor,
            'pedido_id_vincular': pedido_id_vincular
        })
    return redirect('movimientos.index')

@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def movimientos_entrada_procesar(request):
    if request.method == 'POST':
        producto_ids = request.POST.getlist('producto_id[]')
        unidades_empaque = request.POST.getlist('unidad_empaque[]')
        cants_empaques = request.POST.getlist('cant_empaques[]')
        totales = request.POST.getlist('total_unidades[]')
        proveedor_id = request.POST.get('proveedor_id')

        if not proveedor_id:
            messages.error(request, 'Debe seleccionar un proveedor para el ingreso.')
            return redirect('movimientos.entrada')

        proveedor = Proveedor.objects.get(id=proveedor_id)
        ahora = timezone.now()
        lote_id = f"E-{ahora.strftime('%Y%m%d%H%M')}-{str(uuid.uuid4())[:8]}"
        entradas_creadas = 0
        try:
            with transaction.atomic():
                for i in range(len(producto_ids)):
                    producto = Inventario.objects.get(id_producto=producto_ids[i])
                    cant_entrada = int(totales[i])
                    
                    if cant_entrada > 0:
                        # Crear el movimiento REAL de entrada
                        MovimientosInventario.objects.create(
                            producto=producto,
                            tipo_movimiento='ENTRADA',
                            cantidad=cant_entrada,
                            unidad_empaque=unidades_empaque[i],
                            cantidad_empaques=int(cants_empaques[i]),
                            proveedor=proveedor,
                            fecha_movimiento=ahora,
                            codigo_lote=lote_id
                        )
                        # Actualizar stock REAL
                        producto.cantidad += cant_entrada
                        # Asociar proveedor al producto
                        producto.proveedores.add(proveedor)
                        producto.save()
                        entradas_creadas += 1
            
            messages.success(request, f'Se registraron exitosamente {entradas_creadas} ingresos de stock de {proveedor.razonsocial}.')
            
            # AUTOMATIZACIÓN: Cambiar estado del pedido si existe
            pedido_id_input = request.POST.get('pedido_id_vincular')
            if pedido_id_input:
                try:
                    pedido_real = PedidoCompra.objects.get(id_pedido=pedido_id_input)
                    pedido_real.estado = 'COMPLETADO'
                    pedido_real.save()
                    messages.info(request, f"El pedido {pedido_real.codigo_lote} ha sido marcado como COMPLETADO.")
                except PedidoCompra.DoesNotExist:
                    pass

            return redirect('movimientos.index')
            
        except Exception as e:
            messages.error(request, f'Error al procesar las entradas: {str(e)}')
            return redirect('movimientos.entrada')

    return redirect('movimientos.index')

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def compras_editar_pedido(request, pedido_id):
    pedido = get_object_or_404(PedidoCompra, id_pedido=pedido_id)
    proveedor = pedido.proveedor
    # Filtrar productos que tienen asociado este proveedor
    productos = Inventario.objects.filter(proveedores__id=proveedor.id).order_by('nombre_producto')
    
    # Necesitamos las opciones de unidad de empaque para los selectores
    from .models import UnidadEmpaqueChoices
    unidades_choices = UnidadEmpaqueChoices.choices

    # Mapeamos los detalles actuales para facilitar el acceso en el template
    detalles_dict = {d.producto_id: d for d in pedido.detalles.all()}
    
    productos_con_detalles = []
    for p in productos:
        productos_con_detalles.append({
            'producto': p,
            'detalle': detalles_dict.get(p.id_producto)
        })

    return render(request, 'compras/form_pedido.html', {
        'proveedor': proveedor,
        'productos': productos, # se mantiene por si acaso
        'productos_con_detalles': productos_con_detalles,
        'unidades_choices': unidades_choices,
        'pedido': pedido
    })

#======================================
# La decodificación se realiza ahora en el Front-end con Html5-QRCode

#usuarios views
@login_required
# Solo admin puede ver la lista
@user_passes_test(es_admin, login_url='index')
def usuarios_index(request):
    # Parámetros GET
    q = request.GET.get('q', '').strip()
    page_size = 10

    # Usamos el modelo Cliente que ya está importado
    qs = PerfilUsuario.objects.select_related('user').all()

    if q:
        qs = qs.filter(
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q) |
            Q(user__username__icontains=q) |
            Q(direccion__icontains=q) |
            Q(telefono__icontains=q)
        )
    
    paginator = Paginator(qs.order_by('user__first_name'), page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'usuarios/index.html', {'page_obj': page_obj, 'q': q})

@login_required
 # Solo admin puede crear
@user_passes_test(es_admin, login_url='index')
def usuarios_crear(request):
    if request.method == 'POST':
        user_form = UserForm(request.POST)
        perfil_form = PerfilUsuarioForm(request.POST)
        if user_form.is_valid() and perfil_form.is_valid():
            user = user_form.save()
            
            # Al guardar el usuario, la señal ya creó el perfil. Lo recuperamos y actualizamos.
            perfil = user.perfilusuario
            perfil_form = PerfilUsuarioForm(request.POST, instance=perfil)
            
            if perfil_form.is_valid():
                perfil_form.save()
                messages.success(request, 'Usuario creado exitosamente.')
                return redirect('usuarios.index')
    else:
        user_form = UserForm()
        perfil_form = PerfilUsuarioForm()
    return render(request, 'usuarios/crear.html', {'user_form': user_form, 'perfil_form': perfil_form})

@login_required
@user_passes_test(es_admin, login_url='index')
def usuarios_editar(request,id):
    perfil = get_object_or_404(PerfilUsuario, id=id)
    user = perfil.user
    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=user)
        perfil_form = PerfilUsuarioForm(request.POST, instance=perfil)
        if user_form.is_valid() and perfil_form.is_valid():
            user_form.save()
            perfil_form.save()
            messages.success(request, 'Usuario actualizado exitosamente.')
            return redirect('usuarios.index')
    else:
        user_form = UserForm(instance=user)
        perfil_form = PerfilUsuarioForm(instance=perfil)
    return render(request, 'usuarios/editar.html', {'user_form': user_form, 'perfil_form': perfil_form})

@login_required
@user_passes_test(es_admin, login_url='index') # Solo admin puede eliminar
@user_passes_test(es_admin, login_url='index')
def usuarios_eliminar(request,id):
    perfil = get_object_or_404(PerfilUsuario, id=id)
    perfil.user.delete() # Esto elimina el usuario y el perfil en cascada
    messages.success(request, 'Usuario eliminado exitosamente.')
    return redirect('usuarios.index') 

# ======================================
# BACKUPS / COPIAS DE SEGURIDAD
# ======================================

@login_required
@user_passes_test(es_admin, login_url='index')
def realizar_copia_seguridad(request):
    # Buffer para capturar los datos en memoria
    output = io.StringIO()

    # Modelos a incluir en el backup, en formato 'app.Model'
    # Excluimos auth.User y libreria.PerfilUsuario para evitar problemas.
    modelos_a_incluir = [
        'libreria.Proveedor',
        'libreria.Inventario',
        'libreria.MovimientosInventario',
        'libreria.HistorialProveedoresNotas',
    ]

    # Ejecutamos 'dumpdata' para exportar solo los modelos especificados
    call_command('dumpdata', *modelos_a_incluir, stdout=output)
    # Preparamos la respuesta para descargar el archivo
    response = HttpResponse(output.getvalue(), content_type='application/json')
    filename = f"backup_inventario_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.json"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

# ======================================
# PEDIDOS / COMPRAS DE PROVEEDORES
# ======================================

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def compras_seleccionar_proveedor(request):
    q = request.GET.get('q', '').strip()
    qs = Proveedor.objects.all()
    if q:
        qs = qs.filter(Q(razonsocial__icontains=q) | Q(rif__icontains=q))
    
    return render(request, 'compras/seleccionar_proveedor.html', {
        'proveedores': qs,
        'q': q
    })

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def compras_form_pedido(request, proveedor_id):
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)
    # Filtrar productos que tienen asociado este proveedor
    productos = Inventario.objects.filter(proveedores__id=proveedor_id).order_by('nombre_producto')
    
    # Necesitamos las opciones de unidad de empaque para los selectores
    from .models import UnidadEmpaqueChoices
    unidades_choices = UnidadEmpaqueChoices.choices

    return render(request, 'compras/form_pedido.html', {
        'proveedor': proveedor,
        'productos': productos,
        'productos_con_detalles': [{'producto': p, 'detalle': None} for p in productos],
        'unidades_choices': unidades_choices
    })


@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def compras_confirmar(request):
    if request.method == 'POST':
        proveedor_id = request.POST.get('proveedor_id')
        proveedor = get_object_or_404(Proveedor, id=proveedor_id)
        
        producto_ids = request.POST.getlist('producto_id[]')
        items_seleccionados = request.POST.getlist('items_seleccionados[]')
        minimos = request.POST.getlist('minimo[]')
        maximos = request.POST.getlist('maximo[]')
        unidades_empaque = request.POST.getlist('unidad_empaque[]')
        cants_empaques = request.POST.getlist('cant_empaques[]')
        totales = request.POST.getlist('total_unidades[]')
        cants_por_empaque = request.POST.getlist('cant_por_empaque[]')
        pedido_id = request.POST.get('pedido_id')
        
        items_resumen = []
        for i in range(len(producto_ids)):
            p_id = producto_ids[i]
            if p_id in items_seleccionados:
                producto = Inventario.objects.get(id_producto=p_id)
                items_resumen.append({
                    'producto': producto,
                    'min': minimos[i],
                    'max': maximos[i],
                    'unidad': unidades_empaque[i],
                    'cant_empaques': cants_empaques[i],
                    'total': totales[i],
                    'cant_por_empaque': cants_por_empaque[i],
                })
        
        if not items_resumen:
            messages.warning(request, "Debe seleccionar al menos un producto para el pedido.")
            return redirect('compras.nuevo', proveedor_id=proveedor_id)
            
        return render(request, 'compras/confirmar_pedido.html', {
            'proveedor': proveedor,
            'items': items_resumen,
            'pedido_id': pedido_id
        })
    return redirect('compras.index')

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def compras_procesar(request):
    if request.method == 'POST':
        proveedor_id = request.POST.get('proveedor_id')
        proveedor = get_object_or_404(Proveedor, id=proveedor_id)
        
        producto_ids = request.POST.getlist('producto_id[]')
        minimos = request.POST.getlist('minimo[]')
        maximos = request.POST.getlist('maximo[]')
        unidades_empaque = request.POST.getlist('unidad_empaque[]')
        cants_empaques = request.POST.getlist('cant_empaques[]')
        totales = request.POST.getlist('total_unidades[]')
        cants_por_empaque = request.POST.getlist('cant_por_empaque[]')
        pedido_id = request.POST.get('pedido_id')

        ahora = timezone.now()
        # Generar un código más "Socio-Amigable" y único (ORD-AÑO-MES-DIA-RANDOM)
        lote_id = f"ORD-{ahora.strftime('%Y%m%d')}-{str(uuid.uuid4())[:4].upper()}"
        ordenes_creadas = 0
        try:
            with transaction.atomic():
                if pedido_id:
                    pedido = get_object_or_404(PedidoCompra, id_pedido=pedido_id)
                    # Limpiar detalles previos si se está editando
                    pedido.detalles.all().delete()
                    # Opcional: actualizar fecha o mantener la original
                    pedido.fecha_pedido = ahora 
                    pedido.save()
                else:
                    pedido = PedidoCompra.objects.create(
                        proveedor=proveedor,
                        fecha_pedido=ahora,
                        estado='PENDIENTE',
                        codigo_lote=lote_id
                    )
                
                for i in range(len(producto_ids)):
                    producto = Inventario.objects.get(id_producto=producto_ids[i])
                    
                    # (Todo editado a Informativo puro, sin afectar el inventario)
                    
                    # 2. Entrada de Pedido (Informativa)
                    cant_pedir = int(totales[i])
                    if cant_pedir > 0:
                        DetallePedidoCompra.objects.create(
                            pedido=pedido,
                            producto=producto,
                            cantidad=cant_pedir,
                            unidad_empaque=unidades_empaque[i],
                            cantidad_empaques=int(cants_empaques[i]),
                            cantidad_por_empaque=int(cants_por_empaque[i])
                        )
                        ordenes_creadas += 1
            
            messages.success(request, f'Pedido procesado exitosamente. Se registraron {ordenes_creadas} productos en la orden.')
            # Guardamos el ID en la sesión para el prompt de descarga
            request.session['ultimo_pedido_id'] = pedido.id_pedido
            return redirect('movimientos.historial_pedidos')
            
        except Exception as e:
            messages.error(request, f'Error al procesar el pedido: {str(e)}')
            return redirect('compras.nuevo', proveedor_id=proveedor_id)

    return redirect('compras.index')

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def exportar_pedido_unico_pdf(request, pedido_id):
    pedido = get_object_or_404(PedidoCompra, id_pedido=pedido_id)
    
    pdf = FPDF()
    pdf.add_page()
    
    # Cabecera
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 15, "ORDEN DE PEDIDO", ln=True, align="C")
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, f"Código: {pedido.codigo_lote}", ln=True, align="C")
    pdf.ln(5)
    
    # Datos del Proveedor y Fecha
    pdf.set_fill_color(245, 245, 245)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(95, 10, " Proveedor:", 0, 0, "L", True)
    pdf.cell(95, 10, " Detalles:", 0, 1, "L", True)
    
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(95, 7, f" {pedido.proveedor.razonsocial}", 0, 0)
    pdf.cell(95, 7, f" Fecha: {pedido.fecha_pedido.strftime('%d/%m/%Y %H:%M')}", 0, 1)
    pdf.cell(95, 7, f" RIF: {pedido.proveedor.rif}", 0, 0)
    pdf.cell(95, 7, f" Estado: {pedido.estado}", 0, 1)
    pdf.ln(10)
    
    # Tabla de Productos
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(0, 123, 255) # Azul
    pdf.set_text_color(255, 255, 255)
    pdf.cell(80, 10, " Producto", 1, 0, "L", True)
    pdf.cell(35, 10, " Empaques", 1, 0, "C", True)
    pdf.cell(35, 10, " Tipo", 1, 0, "C", True)
    pdf.cell(40, 10, " Total Unid.", 1, 1, "C", True)
    
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 10)
    
    for det in pedido.detalles.all():
        pdf.cell(80, 10, f" {det.producto.nombre_producto[:35]}", 1, 0, "L")
        pdf.cell(35, 10, str(det.cantidad_empaques), 1, 0, "C")
        pdf.cell(35, 10, str(det.unidad_empaque), 1, 0, "C")
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(40, 10, str(det.cantidad), 1, 1, "C")
        pdf.set_font("Helvetica", "", 10)

    pdf.ln(15)
    pdf.set_font("Helvetica", "I", 9)
    pdf.cell(0, 10, "Nota: Este documento es una orden informativa de pedido interna.", ln=True, align="C")
    
    response = HttpResponse(pdf.output(dest='S').encode('latin-1'), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="pedido_{pedido.codigo_lote}.pdf"'
    return response

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def exportar_pedido_pdf(request):
    # Obtener productos con stock <= mínimo
    productos = Inventario.objects.filter(cantidad__lte=F('stock_minimo')).order_by('nombre_producto')
    
    fecha_emision = timezone.now().strftime('%d/%m/%Y %H:%M')
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    
    # Encabezado
    pdf.cell(0, 10, "Reporte de Sugerencia de Pedido (Stock Bajo)", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 10, f"Generado el: {fecha_emision}", ln=True, align="C")
    pdf.ln(5)
    
    if not productos.exists():
        pdf.set_font("Helvetica", "I", 12)
        pdf.cell(0, 20, "No hay productos con stock bajo actualmente.", ln=True, align="C")
    else:
        # Tabla
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_fill_color(220, 53, 69) # Rojo suave para alerta
        pdf.set_text_color(255, 255, 255)
        
        pdf.cell(70, 10, "Producto", 1, 0, "C", True)
        pdf.cell(30, 10, "Stock Act.", 1, 0, "C", True)
        pdf.cell(30, 10, "Stock Min.", 1, 0, "C", True)
        pdf.cell(60, 10, "Pedido Sugerido", 1, 1, "C", True)
        
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(0, 0, 0)
        
        for p in productos:
            sugerido = p.stock_maximo - p.cantidad
            if sugerido < 0: sugerido = 0
            
            pdf.cell(70, 10, p.nombre_producto[:35], 1, 0, "L")
            pdf.cell(30, 10, str(p.cantidad), 1, 0, "C")
            pdf.cell(30, 10, str(p.stock_minimo), 1, 0, "C")
            pdf.cell(60, 10, f"Pedir {sugerido} unid. aprox.", 1, 1, "R")
            
    response = HttpResponse(pdf.output(dest='S').encode('latin-1'), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="pedido_sugerido.pdf"'
    return response

# --- NUEVAS VISTAS DE HISTORIAL AGRUPADO ---

@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def movimientos_historial_pedidos(request):
    q = request.GET.get('q', '').strip()
    qs = PedidoCompra.objects.select_related('proveedor').all()
    
    if q:
        qs = qs.filter(Q(proveedor__razonsocial__icontains=q) | Q(codigo_lote__icontains=q))
        
    qs = qs.prefetch_related('detalles__producto').order_by('-fecha_pedido')
    
    paginator = Paginator(qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Revisar si hay un pedido recién creado para preguntar por su descarga
    ultimo_pedido_id = request.session.pop('ultimo_pedido_id', None)
    
    return render(request, 'compras/historial_pedidos.html', {
        'page_obj': page_obj,
        'titulo': 'Historial de Pedidos de Compra (Informativo)',
        'q': q,
        'modal_pdf_id': ultimo_pedido_id
    })

@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def compras_eliminar_pedido(request):
    if request.method == 'POST':
        pedido_id = request.POST.get('pedido_id')
        try:
            pedido = PedidoCompra.objects.get(id_pedido=pedido_id)
            pedido.delete()
            messages.success(request, 'Pedido eliminado exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al eliminar el pedido: {str(e)}')
            
    return redirect('movimientos.historial_pedidos')

@login_required
@user_passes_test(es_admin, login_url='index')
def compras_eliminar_todo_historial(request):
    if request.method == 'POST':
        try:
            # Eliminar todos los pedidos informativos
            cantidad = PedidoCompra.objects.count()
            PedidoCompra.objects.all().delete()
            messages.success(request, f'Se han eliminado correctamente {cantidad} registros del historial de pedidos.')
        except Exception as e:
            messages.error(request, f'Error al limpiar el historial: {str(e)}')
            
    return redirect('movimientos.historial_pedidos')

@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def movimientos_historial_salidas(request):
    return _historial_agrupado(request, 'SALIDA', 'Historial de Salidas Masivas')

def _historial_agrupado(request, tipo, titulo):
    q = request.GET.get('q', '').strip()
    
    # Obtenemos movimientos filtrados por tipo
    if isinstance(tipo, list):
        qs = MovimientosInventario.objects.filter(tipo_movimiento__in=tipo).select_related('producto', 'proveedor')
        tipo_template = tipo[0]
    else:
        qs = MovimientosInventario.objects.filter(tipo_movimiento=tipo).select_related('producto', 'proveedor')
        tipo_template = tipo
    
    if q:
        qs = qs.filter(
            Q(proveedor__razonsocial__icontains=q)
        )
    
    # Agrupamos por codigo_lote (si existe) o por fecha (exacta)
    # Para los nuevos usamos codigo_lote. Para los viejos, la fecha.
    historial = qs.values('codigo_lote', 'fecha_movimiento', 'proveedor__razonsocial', 'proveedor__rif') \
                  .annotate(
                      total_items=Count('id_movimiento'),
                      total_unidades=Sum('cantidad')
                  ).order_by('-fecha_movimiento')

    lotes = []
    # Procesar los resultados de values() para agrupar realmente por codigo_lote o fecha
    # Usaremos un set para no repetir lotes ya procesados
    lotes_procesados = set()
    
    for grupo in historial:
        # Definir la clave de agrupación: si hay codigo_lote, lo usamos. Si no, la fecha.
        key = grupo['codigo_lote'] if grupo['codigo_lote'] else f"FIXED-{grupo['fecha_movimiento']}-{grupo['proveedor__razonsocial']}"
        
        if key in lotes_procesados:
            continue
            
        if grupo['codigo_lote']:
            detalles = qs.filter(codigo_lote=grupo['codigo_lote'])
        else:
            detalles = qs.filter(
                fecha_movimiento=grupo['fecha_movimiento'],
                proveedor__razonsocial=grupo['proveedor__razonsocial']
            )
            
        lotes.append({
            'info': grupo,
            'detalles': detalles,
            'lote_key': key
        })
        lotes_procesados.add(key)

    # Paginación
    paginator = Paginator(lotes, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'movimientos/historial_lotes.html', {
        'page_obj': page_obj,
        'titulo': titulo,
        'tipo': tipo_template,
        'q': q
    })

@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def exportar_lote_pdf(request):
    fecha_str = request.GET.get('fecha')
    prov_nombre = request.GET.get('prov')
    tipo = request.GET.get('tipo', 'ENTRADA')
    lote_id = request.GET.get('lote')
    
    try:
        if lote_id:
            qs = MovimientosInventario.objects.filter(codigo_lote=lote_id).select_related('producto', 'proveedor')
        else:
            qs = MovimientosInventario.objects.filter(
                fecha_movimiento=fecha_str,
                proveedor__razonsocial=prov_nombre,
                tipo_movimiento=tipo
            ).select_related('producto', 'proveedor')
        
        if not qs.exists():
            messages.error(request, "No se encontró el lote especificado.")
            return redirect('movimientos.index')
            
        # Generar PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        
        titulo_doc = "Comprobante de Pedido / Entrada" if tipo == 'ENTRADA' else "Comprobante de Salida Masiva"
        pdf.cell(0, 10, titulo_doc, ln=True, align="C")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 10, f"Fecha: {qs[0].fecha_movimiento.strftime('%d/%m/%Y %H:%M:%S')}", ln=True, align="C")
        pdf.ln(5)
        
        # Datos del Proveedor / Destino
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 10, "Datos del Proveedor / Destinatario:", ln=True)
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 7, f"Nombre/Razon Social: {qs[0].proveedor.razonsocial}", ln=True)
        pdf.cell(0, 7, f"RIF: {qs[0].proveedor.rif}", ln=True)
        pdf.ln(5)
        
        # Tabla de Productos
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_fill_color(240, 240, 240)
        pdf.cell(80, 10, "Producto", 1, 0, "C", True)
        pdf.cell(40, 10, "Cant. Movida", 1, 0, "C", True)
        pdf.cell(30, 10, "Empaques", 1, 0, "C", True)
        pdf.cell(40, 10, "Unidad", 1, 1, "C", True)
        
        pdf.set_font("Helvetica", "", 10)
        for item in qs:
            pdf.cell(80, 8, item.producto.nombre_producto[:40], 1, 0, "L")
            pdf.cell(40, 8, str(item.cantidad), 1, 0, "C")
            pdf.cell(30, 8, str(item.cantidad_empaques), 1, 0, "C")
            pdf.cell(40, 8, str(item.unidad_empaque), 1, 1, "C")
        
        pdf.ln(10)
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(0, 10, "Documento generado automaticamente por el sistema de inventario.", ln=True, align="R")
        
        response = HttpResponse(pdf.output(dest='S').encode('latin-1'), content_type='application/pdf')
        filename = f"comprobante_{tipo.lower()}_{qs[0].fecha_movimiento.strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        messages.error(request, f"Error al generar el PDF: {str(e)}")
        return redirect('movimientos.index')

@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def movimientos_lote_eliminar(request):
    if request.method == 'POST':
        lote_id = request.POST.get('lote_id')
        fecha_str = request.POST.get('fecha')
        prov_nombre = request.POST.get('prov')
        tipo = request.POST.get('tipo', 'ENTRADA')

        try:
            with transaction.atomic():
                if lote_id:
                    movs = MovimientosInventario.objects.filter(codigo_lote=lote_id).select_related('producto')
                else:
                    movs = MovimientosInventario.objects.filter(
                        fecha_movimiento=fecha_str,
                        proveedor__razonsocial=prov_nombre,
                        tipo_movimiento=tipo
                    ).select_related('producto')

                count = 0
                for mov in movs:
                    producto = mov.producto
                    if mov.tipo_movimiento == 'ENTRADA':
                        producto.cantidad -= mov.cantidad
                    elif mov.tipo_movimiento == 'SALIDA':
                        producto.cantidad += mov.cantidad
                    producto.save()
                    count += 1
                
                movs.delete()
                
            messages.success(request, f'Lote eliminado exitosamente ({count} registros revertidos).')
            
        except Exception as e:
            messages.error(request, f'Error al eliminar el lote: {str(e)}')
            
        return redirect(request.META.get('HTTP_REFERER', 'movimientos.index'))
    return redirect('movimientos.index')

@login_required
@user_passes_test(es_inventario_acceso, login_url='index')
def movimientos_lote_editar(request, lote_id):
    movimientos = MovimientosInventario.objects.filter(codigo_lote=lote_id).select_related('producto', 'proveedor')
    if not movimientos.exists():
        messages.error(request, "Lote no encontrado.")
        return redirect('movimientos.index')

    tipo = movimientos.first().tipo_movimiento

    if request.method == 'POST':
        try:
            with transaction.atomic():
                cambios = 0
                eliminados = 0
                for mov in movimientos:
                    nueva_cantidad = request.POST.get(f'cantidad_{mov.id_movimiento}')
                    nueva_unidad = request.POST.get(f'unidad_{mov.id_movimiento}')
                    
                    if nueva_cantidad and nueva_cantidad.isdigit():
                        nueva_cantidad = int(nueva_cantidad)
                        
                        # Si es 0, eliminamos el registro y revertimos
                        if nueva_cantidad == 0:
                            producto = mov.producto
                            if mov.tipo_movimiento == 'ENTRADA':
                                producto.cantidad -= mov.cantidad
                            elif mov.tipo_movimiento == 'SALIDA':
                                producto.cantidad += mov.cantidad
                            producto.save()
                            mov.delete()
                            eliminados += 1
                            continue

                        # Si cambió algo
                        if nueva_cantidad != mov.cantidad or nueva_unidad != mov.unidad_empaque:
                            producto = mov.producto
                            
                            # Validar stock si es SALIDA y aumentamos
                            if tipo == 'SALIDA' and nueva_cantidad > mov.cantidad:
                                diff = nueva_cantidad - mov.cantidad
                                if producto.cantidad < diff:
                                    raise Exception(f"Stock insuficiente para {producto.nombre_producto}. Faltan {diff - producto.cantidad} unid.")
                            
                            # Revertir
                            if mov.tipo_movimiento == 'ENTRADA':
                                producto.cantidad -= mov.cantidad
                            elif mov.tipo_movimiento == 'SALIDA':
                                producto.cantidad += mov.cantidad
                                
                            # Aplicar nuevo
                            if mov.tipo_movimiento == 'ENTRADA':
                                producto.cantidad += nueva_cantidad
                            elif mov.tipo_movimiento == 'SALIDA':
                                producto.cantidad -= nueva_cantidad
                                
                            producto.save()
                            
                            mov.cantidad = nueva_cantidad
                            mov.cantidad_empaques = nueva_cantidad
                            mov.unidad_empaque = nueva_unidad
                            mov.save()
                            cambios += 1
                            
                # Revisar si se están agregando productos nuevos al lote
                nuevos_prods_ids = request.POST.getlist('nuevo_producto_id[]')
                nuevas_cantidades = request.POST.getlist('nueva_cantidad_prod[]')
                nuevas_unidades = request.POST.getlist('nueva_unidad_prod[]')
                
                nuevos_agregados = 0
                
                # Iterar sobre las listas recibidas para agregar varios productos a la vez
                for i in range(len(nuevos_prods_ids)):
                    nuevo_prod_id = nuevos_prods_ids[i]
                    nueva_cant_prod = nuevas_cantidades[i] if i < len(nuevas_cantidades) else ''
                    nueva_unidad_prod = nuevas_unidades[i] if i < len(nuevas_unidades) else ''
                    
                    if nuevo_prod_id and nueva_cant_prod and str(nueva_cant_prod).isdigit():
                        nueva_cant_prod = int(nueva_cant_prod)
                        if nueva_cant_prod > 0:
                            prod_nuevo = Inventario.objects.get(id_producto=nuevo_prod_id)
                            info = movimientos.first()
                            
                            # Validar stock si es SALIDA
                            if tipo == 'SALIDA' and prod_nuevo.cantidad < nueva_cant_prod:
                                raise Exception(f"Stock insuficiente para {prod_nuevo.nombre_producto}. Disponibles: {prod_nuevo.cantidad}")
                            
                            # Actualizar stock
                            if tipo == 'ENTRADA':
                                prod_nuevo.cantidad += nueva_cant_prod
                            elif tipo == 'SALIDA':
                                prod_nuevo.cantidad -= nueva_cant_prod
                            prod_nuevo.save()
                            
                            # Crear el movimiento
                            MovimientosInventario.objects.create(
                                producto=prod_nuevo,
                                tipo_movimiento=tipo,
                                cantidad=nueva_cant_prod,
                                unidad_empaque=nueva_unidad_prod,
                                cantidad_empaques=nueva_cant_prod,
                                proveedor=info.proveedor,
                                fecha_movimiento=info.fecha_movimiento,
                                codigo_lote=lote_id
                            )
                            nuevos_agregados += 1
                
                msg = f"Lote actualizado correctamente. {cambios} modificados."
                if eliminados > 0: msg += f" {eliminados} eliminados."
                if nuevos_agregados > 0: msg += f" {nuevos_agregados} agregados."
                messages.success(request, msg)
                return redirect('movimientos.historial_pedidos' if tipo == 'ENTRADA' else 'movimientos.historial_salidas')
                
        except Exception as e:
            messages.error(request, f"Error al actualizar lote: {str(e)}")

    from .models import UnidadEmpaqueChoices
    info = movimientos.first()
    
    # Obtener productos disponibles según sea entrada (del proveedor) o salida (todos)
    if tipo == 'ENTRADA':
        # Filtrar por los que sean de ese proveedor
        productos_disponibles = Inventario.objects.filter(proveedores__id=info.proveedor.id).order_by('nombre_producto')
    else:
        # En salida puede usar todos
        productos_disponibles = Inventario.objects.all().order_by('nombre_producto')

    # Excluir los que ya están en el lote
    productos_en_lote = movimientos.values_list('producto_id', flat=True)
    productos_disponibles = productos_disponibles.exclude(id_producto__in=productos_en_lote)

    return render(request, 'movimientos/lote_editar.html', {
        'movimientos': movimientos,
        'lote_id': lote_id,
        'tipo': tipo,
        'info': info,
        'unidades_choices': UnidadEmpaqueChoices.choices,
        'productos_disponibles': productos_disponibles
    })