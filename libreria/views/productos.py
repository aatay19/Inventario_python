import openpyxl
import csv
import io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q, F, Sum
from django.contrib import messages
from django.core.paginator import Paginator
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta
from fpdf import FPDF
from .auth import es_pleno_acceso
from ..models import Inventario, Proveedor
from ..forms import InventarioForm, ImportarArchivoForm

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def inventario_index(request):
    qs = Inventario.objects.all()
    q = request.GET.get('q', '').strip()
    categoria = request.GET.get('categoria', '').strip()
    proveedor_id = request.GET.get('proveedor_id', '').strip()
    order = request.GET.get('order', 'name_asc')
    low = request.GET.get('low', '').strip()
    page_size = 10

    if q:
        qs = qs.filter(
            Q(nombre_producto__icontains=q) |
            Q(codigo_producto__icontains=q)
        )

    if categoria:
        qs = qs.filter(categoria=categoria)
        
    if proveedor_id:
        qs = qs.filter(proveedores__id=proveedor_id)

    if low and low.lower() in ('1', 'true', 'on'):
        qs = qs.filter(cantidad__lte=F('stock_minimo'))

    hace_30_dias = timezone.now() - timedelta(days=30)
    qs = qs.annotate(
        rotacion_30d=Sum(
            'movimientosinventario__cantidad',
            filter=Q(movimientosinventario__tipo_movimiento='SALIDA') & 
                   Q(movimientosinventario__fecha_movimiento__gte=hace_30_dias)
        ),
        total_entradas_hist=Sum(
            'movimientosinventario__cantidad',
            filter=Q(movimientosinventario__tipo_movimiento='ENTRADA')
        ),
        total_salidas_hist=Sum(
            'movimientosinventario__cantidad',
            filter=Q(movimientosinventario__tipo_movimiento='SALIDA')
        )
    )

    if order == 'name_desc':
        qs = qs.order_by('-nombre_producto')
    elif order == 'cantidad_desc':
        qs = qs.order_by('-cantidad')
    elif order == 'cantidad_asc':
        qs = qs.order_by('cantidad')
    elif order == 'rotacion_desc':
        qs = qs.order_by('-rotacion_30d', '-nombre_producto')
    elif order == 'rotacion_asc':
        qs = qs.order_by('rotacion_30d', 'nombre_producto')
    elif order == 'entradas_desc':
        qs = qs.order_by('-total_entradas_hist', '-nombre_producto')
    elif order == 'entradas_asc':
        qs = qs.order_by('total_entradas_hist', 'nombre_producto')
    elif order == 'salidas_desc':
        qs = qs.order_by('-total_salidas_hist', '-nombre_producto')
    elif order == 'salidas_asc':
        qs = qs.order_by('total_salidas_hist', 'nombre_producto')
    else:
        qs = qs.order_by('nombre_producto')

    low_stock_count = qs.filter(cantidad__lte=F('stock_minimo')).count()
    paginator = Paginator(qs, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'q': q,
        'categoria': categoria,
        'order': order,
        'low_stock_count': low_stock_count,
        'low': low,
        'proveedor_id': proveedor_id,
    }
    return render(request, 'inventario/index.html', context)

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def inventario_crear(request):
    formulario_inventario = InventarioForm(request.POST or None)
    if formulario_inventario.is_valid():
        formulario_inventario.save()
        messages.success(request, 'Producto registrado exitosamente.')
        return redirect('/inventario')
    return render(request, 'inventario/crear.html',{'formulario_inventario': formulario_inventario})

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def inventario_editar(request, id_producto):
    producto = Inventario.objects.get(id_producto=id_producto)
    formulario_inventario = InventarioForm(request.POST or None, instance=producto)
    if formulario_inventario.is_valid() and request.POST:
        formulario_inventario.save()
        messages.success(request, 'Producto actualizado exitosamente.')
        return redirect('/inventario')
    return render(request, 'inventario/editar.html',{'formulario_inventario': formulario_inventario})

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def inventario_eliminar(request, id_producto):
    producto = Inventario.objects.get(id_producto=id_producto)
    producto.delete()
    messages.success(request, 'Producto eliminado exitosamente.')
    return redirect('/inventario')

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def inventario_importar(request):
    if request.method == 'POST':
        form = ImportarArchivoForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            nombre_archivo = archivo.name.lower()
            registros_creados = 0
            errores = []

            try:
                datos = []
                if nombre_archivo.endswith('.xlsx'):
                    wb = openpyxl.load_workbook(archivo)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            datos.append(row)
                elif nombre_archivo.endswith('.csv') or nombre_archivo.endswith('.txt'):
                    archivo_data = archivo.read().decode('utf-8')
                    io_string = io.StringIO(archivo_data)
                    reader = csv.reader(io_string, delimiter=',')
                    next(reader, None)
                    for row in reader:
                        if row:
                            datos.append(row)

                for i, fila in enumerate(datos):
                    try:
                        cat_input = str(fila[3]).upper().strip() if len(fila) > 3 else 'OTRO'
                        categoria_valida = 'OTRO'
                        for choice in Inventario._meta.get_field('categoria').choices:
                            if choice[0] == cat_input or choice[1].upper() == cat_input:
                                categoria_valida = choice[0]
                                break
                        
                        obj, created = Inventario.objects.update_or_create(
                            codigo_producto=fila[0],
                            defaults={
                                'nombre_producto': fila[1] if len(fila) > 1 else 'Sin Nombre',
                                'descripcion': fila[2] if len(fila) > 2 else '',
                                'categoria': categoria_valida,
                                'cantidad': int(fila[4]) if len(fila) > 4 and fila[4] else 0,
                                'costo_actual': float(fila[5]) if len(fila) > 5 and fila[5] else 0.0,
                                'stock_minimo': int(fila[6]) if len(fila) > 6 and fila[6] else 5,
                                'stock_maximo': int(fila[7]) if len(fila) > 7 and fila[7] else 100,
                                'unidad_empaque': 'UNIDAD',
                                'cantidad_por_empaque': 1
                            }
                        )

                        if len(fila) > 8 and fila[8]:
                            valor_proveedor = str(fila[8]).strip()
                            proveedor = Proveedor.objects.filter(
                                Q(nombre__icontains=valor_proveedor) | Q(rif__icontains=valor_proveedor)
                            ).first()
                            if proveedor:
                                obj.proveedores.add(proveedor)
                        registros_creados += 1
                    except Exception as e:
                        errores.append(f"Fila {i+2} ({fila[0] if fila else '?'}): {str(e)}")

                messages.success(request, f'Se importaron {registros_creados} productos.')
                if errores:
                    messages.warning(request, f'Errores: {len(errores)}. {errores[0]}')
                return redirect('inventario.index')
            except Exception as e:
                messages.error(request, f'Error general: {str(e)}')
    else:
        form = ImportarArchivoForm()
    return render(request, 'inventario/importar.html', {'form': form})

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def exportar_inventario_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    headers = ["Código", "Producto", "Categoría", "Cantidad", "Unidad Empaque"]
    ws.append(headers)
    productos = Inventario.objects.all()
    for p in productos:
        ws.append([
            p.codigo_producto,
            p.nombre_producto,
            p.get_categoria_display(),
            p.cantidad,
            p.get_unidad_empaque_display()
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario_reporte.xlsx"'
    wb.save(response)
    return response

@login_required
@user_passes_test(es_pleno_acceso, login_url='index')
def exportar_inventario_pdf(request):
    productos = Inventario.objects.all().order_by('nombre_producto')
    total_items = productos.count()
    fecha_emision = timezone.now().strftime('%d/%m/%Y %H:%M')
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Reporte de Inventario General", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 10, f"Fecha de Emision: {fecha_emision}", ln=True, align="C")
    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 10, f"Resumen General:", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 7, f"Total de Productos: {total_items}", ln=True)
    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(30, 10, "Codigo", 1, 0, "C", True)
    pdf.cell(90, 10, "Producto", 1, 0, "C", True)
    pdf.cell(40, 10, "Categoria", 1, 0, "C", True)
    pdf.cell(30, 10, "Stock", 1, 1, "C", True)
    pdf.set_font("Helvetica", "", 9)
    for p in productos:
        nombre = p.nombre_producto[:45]
        categoria = p.get_categoria_display()[:20]
        pdf.cell(30, 8, str(p.codigo_producto), 1, 0, "L")
        pdf.cell(90, 8, nombre, 1, 0, "L")
        pdf.cell(40, 8, categoria, 1, 0, "L")
        pdf.cell(30, 8, str(p.cantidad), 1, 1, "C")
    response = HttpResponse(pdf.output(dest='S').encode('latin-1'), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="inventario_reporte.pdf"'
    return response
@login_required
def buscar_productos_ajax(request):
    from django.http import JsonResponse
    q = request.GET.get('q', '').strip()
    if q:
        qs = Inventario.objects.filter(
            Q(nombre_producto__icontains=q) | Q(codigo_producto__icontains=q)
        )
    else:
        qs = Inventario.objects.all()
    
    # Limitamos a 10 resultados según solicitud del usuario
    productos = qs.order_by('nombre_producto')[:10]
    
    results = [
        {
            'id': p.id_producto,
            'text': f"{p.codigo_producto or 'S/C'} - {p.nombre_producto} (Stock: {p.cantidad})"
        } for p in productos
    ]
    return JsonResponse({'results': results})
